import base64
import re
import csv
import json
import yaml
import pytz
import urllib.request
import decimal
import tempfile
import time
import logging
from types import SimpleNamespace
from decimal import Decimal
import openpyxl
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.writer.excel import save_workbook
import textdistance
import datetime
import charset_normalizer
from io import StringIO, BytesIO
import psycopg2
import pathlib
import os
from trytond.model import ModelSQL, ModelView, fields
from trytond.wizard import Wizard, StateView, StateAction, Button
from trytond.pool import Pool
from trytond.pyson import PYSONEncoder, Eval, Bool
from trytond.transaction import Transaction
from trytond.exceptions import UserError, UserWarning
from trytond.i18n import gettext
from trytond.config import config
from trytond.rpc import RPC
from trytond.report import Report
from trytond.cache import BaseCache
from . import tools

logger = logging.getLogger(__name__)

# As data_to_records will add 'importer_setup' to the context we want to make
# sure it is not used by the cache key as the class is not serializable
# and it would not make sense anyway
BaseCache.context_ignored_keys.add('importer_setup')


DELETES_ENABLED = config.getboolean('importer', 'deletes_enabled', default=False)
DISTANCE_THRESHOLD = config.getfloat('importer', 'distance_threshold',
    default=0.0)
SOFT_LIMIT = 800
LIMIT = 1000

data_sources = [
    ('binary', 'File'),
    ('text', 'Copy & Paste'),
    ('url', 'URL'),
    ('sql', 'SQL'),
    ]

def save_virtual_workbook(workbook):
    with tempfile.NamedTemporaryFile() as tmp:
        save_workbook(workbook, tmp.name)
        with open(tmp.name, 'rb') as f:
            return f.read()

def grouped_slice(records, count=None):
    'grouped_slice implementation that works with iterators'
    if count is None:
        count = Transaction().database.IN_MAX

    chunk = []
    counter = 0
    for record in records:
        chunk.append(record)
        counter += 1
        if counter % count == 0:
            yield chunk
            chunk = []
    if chunk:
        yield chunk


class DataExtractor:
    def __init__(self, data_source, binary_data, text_data, url_data,
            conn=None, sql=None):
        self.data_source = data_source
        self.binary_data = binary_data
        self.text_data = text_data
        self.url_data = url_data
        self.connection = conn
        self.sql = sql
        self.type = None
        self.has_header = None
        self.header_reliable = None
        self.rows = None
        self.filename = None

    @staticmethod
    def to_str(d):
        return charset_normalizer.from_bytes(d).best().output().decode('utf-8')

    def get_data_file(self, force_text=False):
        'force_text makes the method always return a StringIO.'
        'Required by csv reader, which only supports text files.'

        if self.data_source == 'binary' and self.binary_data:
            if force_text:
                return StringIO(self.to_str(self.binary_data))
            return BytesIO(self.binary_data)
        elif self.data_source == 'text' and self.text_data:
            return StringIO(self.text_data)
        elif self.data_source == 'url' and self.url_data:
            url = self.url_data
            if ('docs.google.com' in url and 'export' not in url):
                # Expected URL:
                # https://docs.google.com/spreadsheets/d/19DjZIGvNj1-Z4e4Q4SGqlrnKSAQyMz0JDhhgs2Xbf8w/edit#gid=0
                # New URL:
                # https://docs.google.com/spreadsheets/d/19DjZIGvNj1-Z4e4Q4SGqlrnKSAQyMz0JDhhgs2Xbf8w/export?format=xlsx&id=19DjZIGvNj1-Z4e4Q4SGqlrnKSAQyMz0JDhhgs2Xbf8w&gid=0
                main, _, extra = url.rpartition('/')
                url = main + '/export?format=xlsx&gid=0'
                if '#' in extra:
                    url += '&' + extra.split('#')[-1]
                try:
                    with urllib.request.urlopen(url) as f:
                        data = f.read()
                except:
                    # In some cases we've found that adding gid and other
                    # parameters does not work. In those cases, we try again
                    # without them.
                    url = main + '/export?format=xlsx'
                    with urllib.request.urlopen(url) as f:
                        data = f.read()
            else:
                with urllib.request.urlopen(url) as f:
                    data = f.read()
            if force_text:
                return StringIO(self.to_str(data))
            return BytesIO(data)

        # If no data source or data specified return empty content
        if force_text:
            return StringIO(str())
        return BytesIO(bytes())

    def load(self):
        # Process XLSX files
        try:
            book = openpyxl.load_workbook(filename=self.get_data_file(),
                data_only=True)
        except:
            book = None
        if book:
            sheet = book.active
            rows = []
            for row in sheet.iter_rows():
                # Limit the number of columns to a maximum of 1024.
                # We've found with some spreadsheets with many columns (most of
                # them empty) that not limiting the number of columns causes
                # openpyxl to load data incorrectly. Using 1600 instead of 1024
                # fails too, so we set the limit to 1024 which is the maximum
                # number of columns allowed by LibreOffice
                rows.append([x.value for x in row[:1024]])
            self.type = 'xlsx'
            self.has_header = False
            self.header_reliable = False
            self.rows = rows
            return

        # Process JSON and YAML files
        try:
            content = json.load(self.get_data_file())
            type = 'json'
        except:
            try:
                content = yaml.safe_load(self.get_data_file())
                type = 'yaml'
            except:
                content = None
        if isinstance(content, list):
            if all(isinstance(x, list) for x in content):
                self.type = type
                self.has_header = False
                self.header_reliable = True
                self.rows = content
                return
            if all(isinstance(x, dict) for x in content):
                rows = []
                keys = set()
                # Pick the first 1000 items to determine the header
                for item in content[:1000]:
                    keys.update(item.keys())
                header = [x for x in sorted(list(keys))]
                rows.append(header)
                for item in content:
                    row = []
                    for key in header:
                        row.append(item.get(key))
                    rows.append(row)
                self.type = type
                self.has_header = True
                self.header_reliable = True
                self.rows = rows
                return

        # Process CSV files
        try:
            data = self.get_data_file(force_text=True)
            rows = []
            sniffer = csv.Sniffer()
            chunk = data.read(1024)
            dialect = sniffer.sniff(chunk)
            has_header = sniffer.has_header(chunk)
            data.seek(0)
            reader = csv.reader(data, dialect)
            for row in reader:
                rows.append(row)
            self.type = 'csv'
            self.has_header = has_header
            self.header_reliable = False
            self.rows = rows
            return
        except:
            pass

        if self.connection:
            try:
                cursor = self.connection.cursor()
                cursor.execute(self.sql)
                rows = [[item[0] for item in cursor.description]]
                rows += cursor.fetchall()
                self.type = 'sql'
                self.has_header = True
                self.header_reliable = True
                self.rows = rows
                return
            except Exception:
                pass
        self.type = 'none'
        self.has_header = False
        self.header_reliable = False
        self.rows = []
        return


class Importer(ModelSQL, ModelView):
    'Importer'
    __name__ = 'importer'
    name = fields.Char('Name', required=True)
    method = fields.Selection('get_methods', 'Format', required=True)
    requires_records = fields.Function(fields.Boolean('Requires Records'),
        'get_requires_records')
    language = fields.Many2One('ir.lang', 'Language',
        help='Language to use, if different from the one of the user.')
    model = fields.Function(fields.Many2One('ir.model', 'Model'),
        'on_change_with_model')
    template = fields.Boolean('Template', help="Check to indicate that this "
        "importer is a template and thus should appear in the import wizard.")
    has_header = fields.Boolean('Has Header?')
    use_header = fields.Boolean('Use Header?', states={
            'invisible': ~Eval('has_header'),
            }, help='If checked, the names of the columns will be used for '
            'the mapping. Otherwise, the number of the column will be used.')
    data_source = fields.Selection(
        [(None, ''), ] + data_sources, 'Data Source', states={
            'invisible': ~Eval('data_source_visible'),
            })
    data_source_visible = fields.Function(fields.Boolean('Data Source Visible'),
        'on_change_with_data_source_visible')
    on_error = fields.Selection([
            ('skip', 'Skip'),
            ('raise', 'Raise Error'),
            ('log', 'Log'),
            ], 'On Error', required=True)
    sql_source = fields.Selection([(None, ''), ], 'SQL Source', states={
        'invisible': ~Eval('data_source').in_(['sql']),
        'required': Eval('data_source').in_(['sql']),
        })
    server = fields.Char('Server', states={
        'invisible': ~Eval('data_source').in_(['sql']),
        'required': Eval('data_source').in_(['sql']),
        })
    user = fields.Char('User', states={
        'invisible': ~Eval('data_source').in_(['sql']),
        'required': Eval('data_source').in_(['sql']),
        })
    password = fields.Char('Password', states={
        'invisible': ~Eval('data_source').in_(['sql']),
        'required': Eval('data_source').in_(['sql']),
        })
    database = fields.Char('Database', states={
        'invisible': ~Eval('data_source').in_(['sql']),
        'required': Eval('data_source').in_(['sql']),
        })
    schema = fields.Char('Schema', states={
        'invisible': ~Eval('data_source').in_(['sql']),
        })
    where = fields.Char('Where', states={
        'invisible': ~Eval('data_source').in_(['sql']),
        })
    binary_data = fields.Binary('Data', states={
            'invisible': Eval('data_source') != 'binary',
            }, filename='binary_file_name')
    binary_file_name = fields.Char('Binary File Name', states={
            'invisible': Eval('data_source') != 'binary',
            })
    sql_data = fields.Selection([(None, '')], "SQL Queries", states={
        'invisible': ~Eval('data_source').in_(['sql']),
        'required': Eval('data_source').in_(['sql']),
        })
    text_data = fields.Text('Data', states={
            'invisible': Eval('data_source') != 'text',
            })
    url_data = fields.Char('Data URL', states={
            'invisible': Eval('data_source') != 'url',
            })
    columns = fields.One2Many('importer.column', 'importer', 'Columns')
    source_columns = fields.One2Many('importer.source_column',
            'importer', 'Source Columns', states={
                'invisible': ~Eval('data_source_visible'),
                })
    logs = fields.One2Many('importer.log', 'importer', 'Log')
    errors = fields.Function(fields.One2Many('importer.log', 'importer', 'Errors'), 'get_errors')
    log_success = fields.Boolean('Log Success')
    sample_size = fields.Integer('Sample Size', help="Number of records to "
        "import with the sample button.")
    elapsed = fields.TimeDelta('Elapsed Time', readonly=True)
    deletes = fields.Text('Deletes', readonly=True, states={
            'invisible': ~Bool(Eval('deletes')),
            })

    @staticmethod
    def default_sample_size():
        return 100

    @staticmethod
    def default_language():
        Lang = Pool().get('ir.lang')
        language = Transaction().context.get('language', 'en')
        langs = Lang.search([('code', '=', language)])
        if langs:
            return langs[0].id

    @staticmethod
    def default_on_error():
        return 'skip'

    @staticmethod
    def default_has_header():
        return True

    @staticmethod
    def default_use_header():
        return True

    @classmethod
    def __setup__(cls):
        super().__setup__()
        cls._order.insert(0, ('name', 'ASC'))
        cls._buttons.update({
                'update_columns': {
                    'icon': 'tryton-refresh',
                    'invisible': ~Bool(Eval('requires_records')),
                    'depends': ['requires_records'],
                    },
                'update_source_columns': {
                    'icon': 'tryton-refresh',
                    'invisible': ~Bool(Eval('requires_records')),
                    'depends': ['requires_records'],
                    },
                'detect': {
                    'icon': 'importer-detect',
                    'invisible': ~Bool(Eval('data_source')),
                    },
                'import_': {
                    'icon': 'importer-upload',
                    'invisible': ~Bool(Eval('data_source')) & Bool(Eval('requires_records')),
                    'depends': ['data_source', 'requires_records'],
                    },
                'import_sample': {
                    'icon': 'importer-upload',
                    'invisible': ~Bool(Eval('data_source')),
                    },
                'check_connection': {
                    'icon': 'importer-upload',
                    'invisible': ~Bool(Eval('data_source').in_(['sql'])),
                    },
                'clean_log': {
                    'icon': 'tryton-clear',
                    'invisible': ~Bool(Eval('errors')),
                    },
                'clean_deletes': {
                    'icon': 'tryton-clear',
                    'invisible': ~Bool(Eval('deletes')),
                    },
                'delete_imported': {
                    'icon': 'tryton-cancel',
                    'invisible': ~Bool(Eval('deletes')),
                    },
                })
        cls.sql_source.selection.append(('psql', 'PSQL'))

    def get_language_code(self):
        if self.language:
            return self.language.code
        return Transaction().context.get('language', 'en')

    @classmethod
    def get_errors(cls, importers, name):
        Log = Pool().get('importer.log')

        errors = Log.search([
                ('importer', 'in', importers),
                ('type', '!=', 'success'),
                ])
        res = {x.id: [] for x in importers}
        for error in errors:
            res[error.importer.id].append(error.id)
        return res

    @fields.depends('method')
    def on_change_method(self):
        info = self._get_methods().get(self.method, {})
        if not info.get('requires_records', True):
            self.data_source = None

    @fields.depends('method')
    def on_change_with_data_source_visible(self, name=None):
        info = self._get_methods().get(self.method, {})
        if info.get('requires_records', True):
            return True

    @classmethod
    def create(cls, vlist):
        importers = super().create(vlist)
        return importers

    @classmethod
    def write(cls, *args):
        pool = Pool()
        Warning = pool.get('res.user.warning')

        actions = iter(args)
        for importers, values in zip(actions, actions):
            for importer in importers:
                if 'method' in values and values['method'] != importer.method:
                    key = 'importer_change_method'
                    if Warning.check(key):
                        raise UserWarning(key,
                            gettext('importer.change_method_warning',
                                name=importer.name))
        super().write(*args)

    @classmethod
    def copy(cls, importers, default=None):
        if default is None:
            default = {}
        else:
            default = default.copy()
        default.setdefault('logs')
        default.setdefault('deletes')
        default.setdefault('elapsed')
        return super().copy(importers, default=default)

    @classmethod
    @ModelView.button
    def check_connection(cls, importers):
        for importer in importers:
            if importer.data_source != 'sql':
                continue
            method = getattr(importer, "check_connection_%s" %
                importer.sql_source)
            method()

    def get_connection(self, fail=True):
        if self.data_source != 'sql':
            return
        method = getattr(self, "get_connection_%s" % self.sql_source)
        return method(fail=fail)

    def get_sql(self):
        if self.data_source != 'sql':
            return
        with open(self.get_url_file()) as sql_file:
            sql = sql_file.read()
            sql = sql.format(schema=self.schema, domain=self.where)
            return sql

    def get_url_file(self):

        return os.path.join(pathlib.Path(__file__).parent.resolve(),
            'queries', self.sql_data)

    def check_connection_psql(self):
        try:
            psycopg2.connect(database = self.database, host=self.server,
                user=self.user, password=self.password, port=5432)
        except psycopg2.OperationalError:
            raise UserError(gettext('importer.msg_invalid_connection',
                    importer=self.name))
        raise UserError(gettext('importer.msg_successful_connection',
                importer=self.name))

    def get_connection_psql(self, fail=True):
        try:
            conn = psycopg2.connect(database = self.database, host=self.server,
                user=self.user, password=self.password, port=5432)
        except psycopg2.OperationalError as e:
            if fail:
                raise UserError(e)
            else:
                return None
        return conn


    @classmethod
    def extractor(cls):
        return DataExtractor

    @classmethod
    @ModelView.button
    def clean_log(cls, importers):
        pool = Pool()
        Log = pool.get('importer.log')
        Log.delete(sum([x.logs for x in importers], ()))

    @classmethod
    @ModelView.button
    def clean_deletes(cls, importers):
        for importer in importers:
            importer.deletes = None
        cls.save(importers)

    @classmethod
    @ModelView.button
    def delete_imported(cls, importers):
        if not DELETES_ENABLED:
            raise UserError(gettext('importer.msg_deletes_disabled'))
        cursor = Transaction().connection.cursor()
        for importer in importers:
            if not importer.deletes:
                continue
            cursor.execute(importer.deletes)

    @classmethod
    def _export(cls, importers):
        res = []
        for importer in importers:
            js = {}
            js['name'] = importer.name
            js['method'] = importer.method
            js['language'] = importer.language and importer.language.code
            js['data_source'] = importer.data_source
            js['has_header'] = importer.has_header
            js['use_header'] = importer.use_header
            js['on_error'] = importer.on_error
            # Text
            js['text_data'] = importer.text_data
            # Binary
            if importer.binary_data:
                js['binary_data'] = base64.b64encode(importer.binary_data).decode('utf-8')
            else:
                js['binary_data'] = None
            # URL
            js['url_data'] = importer.url_data
            # SQL
            js['sql_data'] = importer.sql_data
            js['sql_source'] = importer.sql_source
            js['server'] = importer.server
            js['user'] = importer.user
            js['password'] = importer.password
            js['database'] = importer.database
            js['schema'] = importer.schema
            js['where'] = importer.where

            if importer.columns:
                for column in importer.columns:
                    if not column.name and not column.value:
                        continue
                    js['column_field'] = column.field.name
                    js['column_name'] = column.name
                    js['column_value'] = column.value
                    js['column_format'] = column.format
                    res.append(js)
                    js = js.copy()
                    for key in list(js.keys()):
                        js[key] = None
            else:
                res.append(js)
        return json.dumps(res)

    @classmethod
    @ModelView.button
    def update_columns(cls, importers):
        pool = Pool()
        Column = pool.get('importer.column')
        SourceColumn = pool.get('importer.source_column')

        to_delete = []
        to_save = []
        source_to_save = []
        source_to_delete = []
        for importer in importers:
            if not importer.model:
                continue
            # Use the field name so the user can change between
            # similar models without losing the columns
            needed = {x.name: x for x in importer.model.fields}
            for column in importer.columns:
                if column.field.name in needed:
                    column.field = needed[column.field.name]
                    to_save.append(column)
                    del needed[column.field.name]
                else:
                    to_delete.append(column)

            for name, field in needed.items():
                if name in ('id', 'row_number', 'metadata'):
                    continue
                column = Column()
                column.importer = importer
                column.field = field
                to_save.append(column)

            conn = importer.get_connection()
            sql = importer.get_sql()
            Data = cls.extractor()
            data = Data(importer.data_source, importer.binary_data,
                importer.text_data, importer.url_data, conn, sql)
            data.load()

            source_to_delete += importer.source_columns
            if data.rows and importer.has_header:
                # TODO: Source columns should always be created, including
                # when no headers exist. We should use indices instead in those
                # cases.
                importer.generate_source_columns(data.rows[0])
            importer.fill_source_columns()
            SourceColumn.update_examples(importer.source_columns)
            source_to_save += importer.source_columns

        Column.delete(to_delete)
        Column.save(to_save)
        SourceColumn.delete(source_to_delete)
        SourceColumn.save(source_to_save)

    @classmethod
    @ModelView.button
    def update_source_columns(cls, importers):
        pool = Pool()
        Column = pool.get('importer.column')
        SourceColumn = pool.get('importer.source_column')

        columns = []
        to_detect = []
        for importer in importers:
            conn = importer.get_connection()
            sql = importer.get_sql()
            Data = cls.extractor()
            data = Data(importer.data_source, importer.binary_data,
                importer.text_data, importer.url_data, conn, sql)
            data.load()
            if not data.rows:
                continue
            if data.has_header and data.header_reliable:
                importer.has_header = True
                importer.use_header = True
            row = data.rows[0]
            columns += importer.columns
            if not importer.source_columns:
                to_detect.append(importer)
            importer.generate_source_columns(row)
            SourceColumn.update_examples(importer.source_columns)

        cls.save(importers)
        Column.save(columns)
        if to_detect:
            cls.detect(to_detect)

    @classmethod
    @ModelView.button
    def detect(cls, importers, distance_threshold=None):
        pool = Pool()
        Column = pool.get('importer.column')
        SourceColumn = pool.get('importer.source_column')

        columns = []
        source_columns = []
        for importer in importers:
            if not importer.has_header or not importer.use_header:
                continue
            row = [x.name for x in importer.source_columns]
            if not row:
                continue
            columns += importer.columns
            importer.detect_header(row, distance_threshold)
            importer.on_change_columns()
            source_columns += importer.source_columns

        cls.save(importers)
        Column.save(columns)
        SourceColumn.save(source_columns)

    def detect_header(self, row, distance_threshold):
        pool = Pool()
        Lang = pool.get('ir.lang')
        Field = pool.get('ir.model.field')

        if distance_threshold is None:
            distance_threshold = DISTANCE_THRESHOLD

        field_ids = []
        strings = {}
        for column in self.columns:
            field_ids.append(column.field.id)
            strings[column.field] = [column.field.name,
                column.field.field_description]

        langs = Lang.search([
                ('translatable', '=', True),
                ('id', '!=', Transaction().context.get('lang', -1)),
                ])
        for lang in langs:
            with Transaction().set_context(language=lang.code):
                for field in Field.browse(field_ids):
                    strings[field].append(field.field_description)

        lev = textdistance.Levenshtein()
        for column in self.columns:
            row_minimum = (9, None)
            for header in row:
                header = str(header)
                header_minimum = 9
                for string in strings[column.field]:
                    value = lev.normalized_distance(header.lower(),
                        string.lower())
                    header_minimum = min(header_minimum, value)
                if header_minimum < row_minimum[0]:
                    row_minimum = (header_minimum, header)
            if row_minimum[0] <= distance_threshold:
                column.name = row_minimum[1]
            else:
                column.name = None

    def generate_source_columns(self, row):
        pool = Pool()
        SourceColumn = pool.get('importer.source_column')

        self.source_columns = ()
        for pos, field in enumerate(row):
            field = str(field)
            r_column = SourceColumn()
            r_column.importer = self.id
            if self.has_header and self.use_header:
                r_column.name = field
            else:
                r_column.name = str(pos + 1)
            r_column.field = None
            self.source_columns += (r_column,)

    def fill_source_columns(self, name=None):
        pool = Pool()
        Column = pool.get('importer.column')

        for r_column in self.source_columns:
            column = self.column_match_name(r_column.name, Column)
            if column:
                r_column.field = column.field
                r_column.format = column.format

    @fields.depends('columns', 'source_columns', methods=['equal_columns'])
    def on_change_columns(self, name=None):
        for r_column in self.source_columns:
            for column in self.columns:
                if r_column.name == column.name:
                    if self.equal_columns(column, r_column):
                        break
                    r_column.field = column.field
                    r_column.format = column.format
                elif r_column.field == column.field:
                    r_column.field = None

    @fields.depends('columns', 'source_columns', methods=['equal_columns'])
    def on_change_source_columns(self, name=None):
        for column in self.columns:
            for r_column in self.source_columns:
                if column.field == r_column.field:
                    if self.equal_columns(column, r_column):
                        break
                    column.name = r_column.name
                    column.format = r_column.format
                elif column.name == r_column.name:
                    column.name = None

    def column_match_name(self, name, Model):
        columns = Model.search([
            ('importer', '=', self.id),
            ('name', '=', name),
            ], limit=1)
        if columns:
            return columns[0]
        return None

    def equal_columns(self, column1, column2):
        return (column1.name == column2.name and
                column1.field == column2.field and
                column1.format == column2.format)


    @classmethod
    def get_methods(cls):
        return [(k, v['string']) for k, v in cls._get_methods().items()]

    @classmethod
    def _get_methods(cls):
        return {}

    @property
    def chunked(self):
        info = self._get_methods()
        return info[self.method]['chunked']

    def get_requires_records(self, name):
        # Some importers (such as country) don't really require records as
        # input
        info = self._get_methods()
        return info[self.method].get('requires_records', True)

    @fields.depends('method')
    def on_change_with_model(self, name=None):
        pool = Pool()
        Model = pool.get('ir.model')
        item = self._get_methods().get(self.method)
        if not item:
            return
        models = Model.search([('model', '=', item['model'])], limit=1)
        if models:
            return models[0].id

    @classmethod
    @ModelView.button_action('importer.act_import_wizard')
    def import_(cls, importers):
        pass

    @classmethod
    @ModelView.button_action('importer.act_import_sample_wizard')
    def import_sample(cls, importers):
        pass

    def old_data_to_records(self, data=None):
        # Records will be an iterator
        method = getattr(self, 'import_' + self.method)
        if self.requires_records:
            new_records = []
            if self.chunked:
                for records in grouped_slice(self.get_records(data=data)):
                    new_records += method(records)
            else:
                new_records += method(self.get_records(data=data))
        else:
            new_records = method()
        return new_records

    def data_to_records(self, data=None, sample=None):
        pool = Pool()
        Model = pool.get(self.model.model)
        Log = pool.get('importer.log')

        start = time.time()
        if not hasattr(Model, 'importer_start'):
            with Transaction().set_context(language=self.get_language_code()):
                res = self.old_data_to_records(data)
            self.elapsed = datetime.timedelta(seconds=time.time() - start)
            self.save()
            return res

        setup = tools.Setup()
        setup.method = self.method
        setup.on_error = self.on_error
        setup.fields = [x.field.name for x in self.columns if x.name or
            x.value]
        if data:
            setup.filename = data.filename
        else:
            setup.filename = self.binary_file_name
        setup.cache = SimpleNamespace()
        with Transaction().set_context(importer_setup=setup, _no_trigger=True,
                _skip_warnings=True, language=self.get_language_code()):
            Model.importer_start()
            if not self.requires_records:
                return Model.importer_import(fields, [])

            new_records = []
            # In the first iteration we will call the method with a small limit
            # in order to make testing faster.
            soft_limit = SOFT_LIMIT // 10
            limit = LIMIT // 10
            previous_context = {}
            previous_header = None
            subrecords = []
            new_records = []
            count = 0
            batch_start = time.time()
            importer_records = []
            for record in self.get_records(data=data):
                if self.log_success:
                    importer_records.append(record)
                # We do not sort based on context so there can be performance issues
                # if the context changes often
                context = record.importer_context()
                header = record.importer_header(importing=False)
                call = False
                if header is not None:
                    if any(header) and header != previous_header:
                        previous_header = header
                        if (len(subrecords) >= soft_limit
                                or context != previous_context):
                            call = True
                elif len(subrecords) >= limit:
                    call = True
                elif context != previous_context:
                    call = True

                if call:
                    with Transaction().set_context(previous_context):
                        batch = Model.importer_import(subrecords)
                    new_records += batch
                    logger.info('Batch (imported/processed/time): %d/%d/%.2f. '
                        'Total (imported/processed/time): %d/%d/%.2f.',
                        len(batch), len(subrecords), time.time() - batch_start,
                        len(new_records), count, time.time() - start)
                    subrecords = []
                    batch_start = time.time()
                    call = False
                    soft_limit = SOFT_LIMIT
                    limit = LIMIT
                count += 1
                subrecords.append(record)
                previous_context = context
                if sample and count >= sample:
                    break

            if subrecords:
                with Transaction().set_context(previous_context):
                    batch = Model.importer_import(subrecords)
                new_records += batch
                logger.info('Batch (imported/processed/time): %d/%d/%.2f. '
                    'Total (imported/processed/time): %d/%d/%.2f.',
                    len(batch), len(subrecords), time.time() - batch_start,
                    len(new_records), count, time.time() - start)

        if self.logs:
            Log.delete(self.logs)

        importer_records = set(importer_records)
        if setup.errors:
            to_save = []
            generics = set()
            for record, message, kwargs in setup.errors[:setup.limit]:
                generics.add(message)
                log = Log()
                log.importer = self
                if kwargs:
                    log.message = message % kwargs
                else:
                    log.message = message
                if record:
                    log.row_number = record.row_number
                    log.metadata = record.metadata
                    log.source_record = record.to_str(fields=setup.fields)
                    importer_records.discard(record)
                log.type = 'specific'
                to_save.append(log)
            for message in generics:
                log = Log()
                log.importer = self
                log.message = message
                log.type = 'generic'
                to_save.insert(0, log)
            Log.save(to_save)
        if self.log_success:
            to_save = []
            for model, pairs in setup._saved.items():
                for pair in pairs:
                    record_id = pair[0]
                    importer_record = pair[1]
                    importer_records.discard(importer_record)

                    log = Log()
                    log.importer = self
                    if importer_record:
                        log.row_number = importer_record.row_number
                        log.metadata = importer_record.metadata
                        log.source_record = importer_record.to_str(
                            fields=setup.fields)
                    log.record = (model, record_id)
                    log.type = 'success'
                    to_save.append(log)

            for importer_record in importer_records:
                log = Log()
                log.importer = self
                log.row_number = importer_record.row_number
                log.metadata = importer_record.metadata
                log.source_record = importer_record.to_str(fields=setup.fields)
                log.type = 'success'
                to_save.append(log)
            Log.save(to_save)

        self.elapsed = datetime.timedelta(seconds=time.time() - start)
        self.deletes = setup.deletes()
        self.save()
        return new_records

    def get_records(self, raise_errors=True, data=None):
        pool = Pool()
        methods = self._get_methods()
        Model = pool.get(methods[self.method]['model'])

        if data is None:
            conn = self.get_connection()
            sql = self.get_sql()
            Data = self.extractor()
            data = Data(self.data_source, self.binary_data, self.text_data,
                self.url_data, conn, sql)
            data.filename = self.binary_file_name
            data.load()
        rows = data.rows
        if not rows:
            return []
        indexes = self.get_field_indexes(rows)
        if self.has_header:
            rows = rows[1:]

        # We want to make sure we set all fields, even if the Importer record
        # has not been updated since the last change of the model
        missing_fields = ({f.name for f in self.model.fields}
            - {c.field.name for c in self.columns})

        row_number = 0
        if hasattr(Model, 'row_number'):
            update_row_number = True
        else:
            update_row_number = False
        for row in rows:
            row_number += 1
            if not any(row):
                continue
            record = Model()
            # Loop on columns so we ensure we set a value for all fields
            # hence importer methods can rely on the field to exist even if it
            # is None
            for column in self.columns:
                index = indexes.get(column.field.name)
                if index is None:
                    value = column.cast_value(column.value)
                else:
                    try:
                        value = column.cast_value(row[index])
                    except IndexError:
                        if raise_errors:
                            raise UserError(gettext('importer.invalid_index',
                                    column=column.rec_name,
                                    importer=self.rec_name))
                        else:
                            value = None
                    except Exception as e:
                        if raise_errors:
                            raise UserError(gettext('importer.msg_invalid_value',
                                    row=row_number,
                                    column=column.rec_name,
                                    importer=self.rec_name,
                                    value=row[index],
                                    error=e))
                        else:
                            value = None
                    if value is None and column.value:
                        value = column.cast_value(column.value)
                try:
                    setattr(record, column.field.name, value)
                except TypeError:
                    pass

            for field in missing_fields:
                setattr(record, field, None)

            if update_row_number:
                record.row_number = row_number
                record.metadata = None

            yield record

    def get_field_indexes(self, rows):
        indexes = {}
        if self.has_header and self.use_header:
            header = rows[0]
            hi = {}
            for pos in range(len(header)):
                hi[header[pos]] = pos

            for column in self.columns:
                if not column.name:
                    continue
                index = hi.get(column.name)
                if index is None:
                    continue
                indexes[column.field.name] = index
        else:
            indexes = {}
            for column in self.columns:
                if not column.name:
                    continue
                try:
                    index = int(column.name) - 1
                except ValueError:
                    # Convert column name to index
                    index = 0
                    for letter in column.name.upper():
                        index *= 26
                        index += ord(letter) - ord('A') + 1
                    index -= 1
                indexes[column.field.name] = index
        return indexes

    @classmethod
    def datetime_to_utc(cls, datetime_, timezone=None):
        pool = Pool()
        Company = pool.get('company.company')
        company = Company(Transaction().context.get('company'))

        if company.timezone and isinstance(datetime_, datetime.date):
            timezone = pytz.timezone(company.timezone)
            company_datetime = timezone.localize(datetime_, is_dst=None)
            datetime_ = company_datetime.astimezone(pytz.utc)
        return datetime_


class ImporterColumn(ModelSQL, ModelView):
    'Importer Column'
    __name__ = 'importer.column'

    importer = fields.Many2One('importer', 'Importation', required=True,
        ondelete='CASCADE')
    model = fields.Function(fields.Many2One('ir.model', 'Model'),
        'on_change_with_model')
    field = fields.Many2One('ir.model.field', 'Field', ondelete='CASCADE',
        required=True, readonly=True, domain=[
            ('model_ref', '=',
                Eval('_parent_importer', Eval('context', {})).get('model', -1))
        ])
    name = fields.Char('Column Name')
    value = fields.Char('Value', help="Value to be used if no column is "
        "specified or the value found is None.")
    format = fields.Selection('_get_formats', 'Format')
    examples = fields.Function(fields.Char('Examples'),
        'get_examples')

    @fields.depends('importer', '_parent_importer.model')
    def on_change_with_model(self, name=None):
        if self.importer and self.importer.model:
            return self.importer.model.id

    @classmethod
    def __setup__(cls):
        super().__setup__()
        cls._order.insert(0, ('field.field_description', 'ASC'))
        cls.__rpc__.update(
            autocomplete_name=RPC(instantiate=0),
            )

    @classmethod
    def _get_formats(cls):
        return [
            (None, ''),
            ('keep-spaces', 'Keep Spaces'),
            ('date-%d/%m/%Y %H:%M:%S', '%d/%m/%Y %H:%M:%S'),
            ('date-%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M:%S'),
            ('date-%d/%m/%Y', '%d/%m/%Y'),
            ('date-%Y-%m-%d', '%Y-%m-%d'),
            ('decimal-,', 'Decimal (,)'),
            ('decimal-.', 'Decimal (.)'),
            ("decimal-'", "Decimal (')"),
            ]

    @fields.depends('importer', '_parent_importer.source_columns')
    def autocomplete_name(self):
        if not self.importer:
            return []
        return sorted([x.name for x in self.importer.source_columns])

    @classmethod
    def get_examples(self, columns, name):
        res = {x.id: None for x in columns}
        for column in columns:
            if not column.name:
                continue
            for sc in column.importer.source_columns:
                if sc.name == column.name:
                    res[column.id] = sc.examples
                    break
        return res

    def get_selection_dict(self, model, field):
        Lang = Pool().get('ir.lang')
        Model = Pool().get(model)

        d = {}

        for language in Lang.search([('translatable', '=', True)]):
            with Transaction().set_context(language=language.code):
                for selection in Model.fields_get([field])[field]["selection"]:
                    if selection[0] not in d:
                        d[selection[0]] = selection[0]
                    if selection[1] not in d:
                        d[selection[1]] = selection[0]
        return d


    def cast_value(self, value):
        if value is None:
            return value
        ttype = self.field.ttype
        help = self.field.help

        if ttype in ('char', 'text'):
            if help.startswith('selection|'):
                _, model, field = help.split('|')
                d = self.get_selection_dict(model, field)
                if value.strip() not in d:
                    raise UserError(gettext(
                        'importer.value_not_in_selection', field=field))
                return d[value.strip()]
            if isinstance(value, str):
                if self.format and self.format == 'keep-spaces':
                    return value
                return value.strip()
            elif isinstance(value, (float, Decimal)):
                value = str(value)
                if value.endswith('.0'):
                    value = value[:-2]
            else:
                return str(value)
        elif ttype in ('integer', 'float', 'numeric'):
            if isinstance(value, str):
                if self.format and self.format.startswith('decimal-'):
                    decimal_symbol = self.format[len('decimal-'):]
                    for symbol in " ,.'":
                        if symbol == decimal_symbol:
                            continue
                        value = value.replace(symbol, '')
                    value = value.replace(decimal_symbol, '.')
            try:
                pool = Pool()
                ModelClass = pool.get(self.field.model)
                field = getattr(ModelClass, self.field.name)
                if ttype == 'float':
                    value = float(value)
                    if field.digits:
                        value = round(value, field.digits[1])
                    return value
                elif ttype == 'integer':
                    return int(value)
                elif ttype == 'numeric':
                    if isinstance(value, float):
                        value = Decimal('%.10f' % value)
                    else:
                        value = Decimal(value)
                    if field.digits:
                        value = value.quantize(Decimal(str(
                                    10 ** -field.digits[1])))
                    return value
            except (ValueError, decimal.InvalidOperation):
                # TODO: Raise Error
                return None
        elif ttype == 'date':
            if isinstance(value, str):
                value = value.strip()
                if self.format and self.format.startswith('date-'):
                    format = self.format[len('date-'):]
                else:
                    format = '%Y-%m-%d'
                try:
                    return datetime.datetime.strptime(value, format).date()
                except ValueError:
                    # TODO: Raise Error
                    return None
            elif isinstance(value, datetime.datetime):
                return value.date()
            elif not isinstance(value, datetime.date):
                return None
        elif ttype in ('datetime', 'timestamp'):
            if isinstance(value, str):
                value = value.strip()
                if self.format and self.format.startswith('date-'):
                    format = self.format[len('date-'):]
                else:
                    format = '%Y-%m-%d'
                try:
                    return datetime.datetime.strptime(value, format)
                except ValueError:
                    # TODO: Raise Error
                    return None
            elif not isinstance(value, (datetime.datetime, datetime.date)):
                return None
        elif ttype == 'boolean':
            value = str(value).strip()
            if value.lower() in ('false', 'off', '0', ''):
                return False
            else:
                return True
        return value

    @classmethod
    def validate(cls, columns):
        for column in columns:
            column.check_name()

    def check_name(self):
        if not self.name or self.importer.use_header:
            return
        try:
            int(self.name)
        except ValueError:
            if not re.fullmatch('[A-Z]+', self.name):
                raise UserError(gettext('importer.invalid_column_name',
                        name=self.name, column=self.rec_name,
                        importer=self.importer))


class ImporterSourceColumn(ModelSQL, ModelView):
    'Importer Source Column'
    __name__ = 'importer.source_column'

    importer = fields.Many2One('importer', 'Importation', required=True,
        ondelete='CASCADE')
    name = fields.Char('Column Name', required=True, readonly=True)
    model = fields.Function(fields.Many2One('ir.model', 'Model'),
        'on_change_with_model')
    field = fields.Many2One('ir.model.field', 'Field',
            domain=[('model_ref', '=', Eval('model')),])
    format = fields.Selection('_get_formats', 'Format')
    examples = fields.Char('Examples', readonly=True)

    @fields.depends('importer', '_parent_importer.model')
    def on_change_with_model(self, name=None):
        if self.importer and self.importer.model:
            return self.importer.model.id

    @classmethod
    def _get_formats(cls):
        pool = Pool()
        Column = pool.get('importer.column')
        return Column._get_formats()


    @classmethod
    def update_examples(cls, columns):
        pool = Pool()
        Lang = pool.get('ir.lang')

        language = Transaction().context.get('language') or 'en'
        language, = Lang.search([('code', '=', language)])

        def value_to_str(value):
            if isinstance(value, str):
                return value
            if isinstance(value, datetime.date):
                return language.strftime(value)
            return str(value)

        importers = {}
        for column in columns:
            importer = column.importer
            if importer in importers:
                continue
            conn = importer.get_connection()
            sql = importer.get_sql()
            Data = column.importer.extractor()
            data = Data(importer.data_source, importer.binary_data,
                importer.text_data, importer.url_data, conn, sql)
            data.load()
            records = {}
            if data.rows:
                headers = data.rows[0]
                rows = data.rows[1:]
                for i in range(min(len(rows), 3)):
                    row = rows[i]
                    for j in range(len(headers)):
                        records.setdefault(headers[j], []).append(row[j])
            importers[column.importer] = records

        for column in columns:
            records = importers[column.importer]
            try:
                values = records[column.name]
            except KeyError:
                values = []
            if all([x is None for x in values]):
                column.examples = None
            else:
                column.examples = ' | '.join([value_to_str(x) for x in values])


class ImporterLog(ModelSQL, ModelView):
    'Importer Log'
    __name__ = 'importer.log'
    _rec_name = 'message'
    importer = fields.Many2One('importer', 'Importer', required=True,
        readonly=True, ondelete='CASCADE')
    message = fields.Text('Message', readonly=True, states={
            'required': Eval('type') != 'success',
            })
    metadata = fields.Text('Metadata', readonly=True)
    row_number = fields.Integer('Row Number', readonly=True, states={
            'invisible': Eval('type') == 'generic',
            })
    source_record = fields.Text('Source Record', readonly=True, states={
            'invisible': Eval('type') == 'generic',
            })
    record = fields.Reference('Record', selection='get_models', readonly=True,
        states={
            'invisible': Eval('type') != 'success',
            })
    on_error = fields.Selection([
            (None, ''),
            ('skip', 'Skip'),
            ('raise', 'Raise Error'),
            ('log', 'Log'),
            ], 'On Error', states={
            'invisible': Eval('type') == 'success',
            'required': Eval('type') != 'success',
            })
    type = fields.Selection([
            ('success', 'Success'),
            ('generic', 'Generic'),
            ('specific', 'Specific'),
            ], 'Type', required=True, readonly=True)

    @staticmethod
    def default_on_error():
        return 'skip'

    @classmethod
    def __setup__(cls):
        super().__setup__()
        cls._order.insert(0, ('importer', 'ASC'))
        cls._order.insert(1, ('row_number', 'ASC'))

    @staticmethod
    def get_models():
        pool = Pool()
        Model = pool.get('ir.model')
        ModelAccess = pool.get('ir.model.access')
        models = Model.get_name_items()
        if Transaction().check_access:
            access = ModelAccess.get_access([m for m, _ in models])
            models = [(m, n) for m, n in models if access[m]['read']]
        return [(None, '')] + models


class ImportAsk(ModelView):
    'Import Ask'
    __name__ = 'importer.import.ask'
    importer = fields.Many2One('importer', 'Importer', required=True, domain=[
            ('template', '=', True),
            ])
    data_source = fields.Selection(data_sources, 'Data Source', required=True)
    binary_data = fields.Binary('Data', states={
            'invisible': Eval('data_source') != 'binary',
            }, filename='filename')
    filename = fields.Char('Filename')
    text_data = fields.Text('Data', states={
            'invisible': Eval('data_source') != 'text',
            })
    url_data = fields.Char('Data URL', states={
            'invisible': Eval('data_source') != 'url',
            })


class AskAndImportError(ModelView):
    'Import Ask and Import Error'
    __name__ = 'importer.import.ask.error'


class AskAndImport(Wizard):
    'Import'
    __name__ = 'importer.ask_and_import'
    start_state = 'ask'
    ask = StateView('importer.import.ask', 'importer.import_ask_view_form', [
            Button('Cancel', 'end', 'tryton-cancel'),
            Button('Import', 'import_', 'importer-upload', default=True),
            ])
    import_ = StateAction('importer.act_import_open')
    errors = StateAction('importer.act_import_log')
    error_form = StateView('importer.import.ask.error', 'importer.import_ask_error_view_form', [
            Button('Ok', 'end', 'tryton-ok'),
            ])

    def do_import_(self, action):
        Importer = Pool().get('importer')

        conn = self.ask.importer.get_connection()
        sql = self.ask.importer.get_sql()

        Data = Importer.extractor()
        data = Data(self.ask.data_source, self.ask.binary_data,
            self.ask.text_data, self.ask.url_data, conn, sql)
        data.filename = getattr(self.ask, 'filename', None)
        data.load()
        records = self.ask.importer.data_to_records(data)
        if not records:
            Transaction().commit()
            raise UserError(gettext('importer.no_records_imported',
                importer=self.ask.importer.rec_name))

        # TODO: One tab for each model
        models = {}
        for record in records:
            models.setdefault(record.__name__, []).append(record)

        for model, records in models.items():
            action['res_model'] = model
            action['pyson_domain'] = PYSONEncoder().encode(
                [('id', 'in', [x.id for x in records])],
                )
        return action, {}

    def transition_import_(self):
        if self.ask.importer.errors:
            return 'errors'
        return 'end'

    def transition_errors(self):
        return 'error_form'

    def default_errors(self, data):
        return {
            'errors': [x.id for x in self.ask.importer.errors],
            }

    def do_errors(self, action):
        action['pyson_domain'] = PYSONEncoder().encode(
                [('importer', '=', self.ask.importer.id)],
                )
        return action, {}


class Import(Wizard):
    'Import'
    __name__ = 'importer.import'
    start_state = 'import_'
    import_ = StateAction('importer.act_import_open')

    def do_import_(self, action):
        # TODO: Support importing several importers at once
        records = self.record.data_to_records()
        if not records:
            Transaction().commit()
            raise UserError(gettext('importer.no_records_imported',
                importer=self.record.rec_name))

        models = {}
        for record in records:
            models.setdefault(record.__name__, []).append(record)

        for model, records in models.items():
            action['res_model'] = model
            action['pyson_domain'] = PYSONEncoder().encode(
                [('id', 'in', [x.id for x in records])],
                )
        return action, {}

    def transition_import_(self):
        return 'end'


class ImportSample(Wizard):
    'Import Sample'
    __name__ = 'importer.import_sample'
    start_state = 'import_'
    import_ = StateAction('importer.act_import_open')

    def do_import_(self, action):
        records = self.record.data_to_records(sample=self.record.sample_size or 100)

        models = {}
        for record in records:
            models.setdefault(record.__name__, []).append(record)

        for model, records in models.items():
            action['res_model'] = model
            action['pyson_domain'] = PYSONEncoder().encode(
                [('id', 'in', [x.id for x in records])],
                )
        return action, {}

    def transition_import_(self):
        return 'end'


class ExcelTemplate(Report):
    'Excel Template'
    __name__ = 'importer.excel.template'

    @classmethod
    def execute(cls, ids, data):
        if not ids:
            return
        cls.check_access()
        pool = Pool()
        Importer = pool.get('importer')
        importer = Importer(ids[0])
        wb = Workbook()
        ws = wb.active
        header = []
        for column in importer.columns:
            header.append(column.field.field_description)
        header = tuple(header)
        ws.append(header)

        for column in importer.columns:
            if column.field.help and column.field.help.startswith("selection"):
                _, model, field_name = column.field.help.split("|")
                ModelClass = pool.get(model)
                selections = ModelClass.fields_get([field_name])[field_name]['selection']
                selections = [i[0] for i in selections]
                c = 2
                for selection in selections:
                    ws.cell(row=c, column=importer.columns.index(column)+1).value = selection
                    c+= 1

        for column, number in zip(importer.columns, range(1, len(header) + 1)):
            c = column.field.help
            if c:
                ws.cell(row=1, column=number).comment = Comment(c, "Tryton")
        return ('xlsx', save_virtual_workbook(wb), False, importer.name)


class Export(Report):
    'Export'
    __name__ = 'importer.export'

    @classmethod
    def execute(cls, ids, data):
        if not ids:
            return
        Importer = Pool().get('importer')
        importers = Importer.browse(ids)
        return ('json', Importer._export(importers), False, 'importers.json')
